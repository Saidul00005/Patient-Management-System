from openpyxl import Workbook
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
import re
from django.views.decorators.csrf import csrf_exempt
from django.conf import settings
from django.http import HttpResponse
from authentication.models import CustomUser
from .serializers import AdminSerializer, CustomUserSerializer
from authentication.models import Admin
from .models import Client, Staff, Appointment, AppointmentSendInfo, SendType
from .serializers import ClientSerializer, StaffSerializer
from django.conf import settings
from .filters import UserSearchFilter

###############################33
from .models import Appointment
from .serializers import *  # Assuming you have a serializer for Appointment model
from datetime import datetime
from authentication.custom_auth import IsAdminOrStaff, IsAdmin
from django.utils.timezone import now
from datetime import timedelta
import requests
from django.shortcuts import render
from rest_framework import generics
from django.http import HttpResponse
from openpyxl import Workbook
from io import BytesIO
from datetime import datetime
from datetime import datetime
from .models import Company
from rest_framework import generics, status, serializers

##########################################15/07/24 FOR DELEDE USER and EDIT USER
from rest_framework.permissions import IsAdminUser
from django.contrib.auth import get_user_model
from rest_framework.exceptions import NotFound
from django.utils import timezone

from rest_framework.generics import ListAPIView
from authentication.models import Staff, Client, ClientComment
from authentication.serializers import (
    StaffUserSerializer,
    ClientUserSerializer,
    ClientCommentSerializer,
)

User = get_user_model()

######################################
# from .permissions import IsAdminOrStaff  # Ensure this is defined in your permissions
# from .filters import UserSearchFilter  # Ensure this is your custom filter


class CompanyView(APIView):
    permission_classes = [
        AllowAny,
    ]
    serializer_class = CompanySerializer

    def get(self, request):
        company = Company.objects.all()
        # the many param informs the serializer that it will be serializing more than a single article.
        serializer = CompanySerializer(company, many=True)
        return Response({"Company": serializer.data})


class CompanyUpdateView(generics.RetrieveUpdateDestroyAPIView):
    serializer_class = CompanySerializer
    permission_classes = [
        AllowAny,
    ]
    lookup_field = "id"

    def get_queryset(self):
        return Company.objects.all()


def home(request):
    return HttpResponse("Hello World!")


class Login(APIView):
    permission_classes = [
        AllowAny,
    ]
    serializer_class = LoginSerializer

    def post(self, request):
        login_serializer = self.serializer_class(data=request.data)
        if login_serializer.is_valid():
            data = login_serializer.save()
            return Response(data, status=status.HTTP_200_OK)
        else:
            return Response(login_serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class CreateCompany(APIView):
    permission_classes = [
        IsAdmin,
    ]
    serializer_class = CompanySerializer

    def post(self, request):
        company_serializer = self.serializer_class(data=request.data)
        if company_serializer.is_valid():
            company_serializer.save()
            return Response(
                "Company created successfully", status=status.HTTP_201_CREATED
            )
        else:
            return Response(
                company_serializer.errors, status=status.HTTP_400_BAD_REQUEST
            )


# class CreateUserView(APIView):
#     permission_classes = [IsAdminOrStaff]
#     serializer_class = CustomUserSerializer

#     def post(self, request):
#         serializer = CustomUserSerializer(data = request.data,context = { 'user': request.user})

#         if serializer.is_valid():
#             user = serializer.save()
#             msg = f"{user.user_role.capitalize()} user created successfully."
#             res = {'msg': msg}
#             return Response(res, status=status.HTTP_201_CREATED)
#         else:
#             return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class UserEditView(APIView):
    permission_classes = [IsAdminOrStaff]
    serializer_class = UserEditSerializer

    def edit_user(self, request, pk, partial):
        try:
            user = User.objects.get(pk=pk)
        except User.DoesNotExist:
            return Response(
                {"error": "User not found"}, status=status.HTTP_404_NOT_FOUND
            )

        self.check_object_permissions(request, user)
        print(request.data)
        serializer = self.serializer_class(
            user, data=request.data, partial=True, context={"user": request.user}
        )
        if serializer.is_valid():
            user = serializer.save()
            msg = f"{user.user_role.capitalize()} user updated successfully."
            res = {"msg": msg}
            return Response(res, status=status.HTTP_200_OK)
        else:
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def patch(self, request, pk):
        return self.edit_user(request, pk, partial=True)

    def put(self, request, pk):
        return self.edit_user(request, pk, partial=False)


class UserView(APIView):
    permission_classes = [IsAdminOrStaff]
    serializer_class = CustomUserSerializer

    def post(self, request):
        serializer = self.serializer_class(
            data=request.data, context={"user": request.user}
        )
        if serializer.is_valid():
            user = serializer.save()
            msg = f"{user.user_role.capitalize()} user created successfully."
            res = {"msg": msg}
            return Response(res, status=status.HTTP_201_CREATED)
        else:
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def patch(self, request, pk):
        try:
            user = User.objects.get(pk=pk)
        except User.DoesNotExist:
            return Response(
                {"error": "User not found"}, status=status.HTTP_404_NOT_FOUND
            )

        self.check_object_permissions(request, user)

        serializer = self.serializer_class(
            user, data=request.data, partial=True, context={"user": request.user}
        )
        if serializer.is_valid():
            user = serializer.save()
            msg = f"{user.user_role.capitalize()} user updated successfully."
            res = {"msg": msg}
            return Response(res, status=status.HTTP_200_OK)
        else:
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def delete(self, request, pk):
        try:
            user = User.objects.get(pk=pk)
        except User.DoesNotExist:
            return Response(
                {"error": "User not found"}, status=status.HTTP_404_NOT_FOUND
            )

        self.check_object_permissions(request, user)

        user.delete()
        return Response(
            {"message": "User deleted successfully"}, status=status.HTTP_204_NO_CONTENT
        )

    def get(self, request, pk):
        try:
            user = User.objects.get(pk=pk)
        except User.DoesNotExist:
            return Response(
                {"error": "User not found"}, status=status.HTTP_404_NOT_FOUND
            )

        serializer = self.serializer_class(user)
        return Response(serializer.data, status=status.HTTP_200_OK)


class GetProfile(APIView):
    permission_classes = [IsAuthenticated]
    serializer_class = CustomUserSerializer

    def get(self, request):
        serializer = self.serializer_class(request.user, many=False)
        return Response(serializer.data, status=status.HTTP_200_OK)


###########################################################################15/07/24
# class DeleteUserView(APIView):
#     permission_classes = [IsAdminUser]

#     def delete(self, request, user_id):
#         try:
#             user = User.objects.get(id=user_id)
#         except User.DoesNotExist:
#             raise NotFound(detail="User not found.")

#         user.delete()
#         return Response({"msg": "User deleted successfully."}, status=status.HTTP_204_NO_CONTENT)

# class DeleteUserView(APIView):
#     permission_classes = [IsAdminUser]

#     def delete(self, request, user_id):
#         try:
#             user = User.objects.get(id=user_id)
#         except User.DoesNotExist:
#             raise NotFound(detail="User not found.")

#         # Print user details before deleting
#         user_details = {
#             "id": user.id,
#             "username": user.username,
#             "email": user.email,
#             # Add other details as needed
#         }
#         print(f"User to be deleted: {user_details}")

#         # Delete the user
#         user.delete()
#         return Response({"msg": "User deleted successfully."}, status=status.HTTP_204_NO_CONTENT)
#
#
#
# DELETE USER FUNCTION
import logging

logger = logging.getLogger("django")


class DeleteUserView(APIView):
    permission_classes = [IsAdminUser]

    def delete(self, request, user_id):
        try:
            user = User.objects.get(id=user_id)
        except User.DoesNotExist:
            raise NotFound(detail="User not found.")

        # Determine user type
        user_type = "Unknown"
        if hasattr(user, "staff"):
            user_type = "Staff"
        elif hasattr(user, "client"):
            user_type = "Client"
        elif hasattr(user, "admin"):
            user_type = "Admin"

        # Log user details before deleting
        user_details = {
            "id": user.id,
            "username": user.username,
            "email": user.email,
            "type": user_type,
            # Add other details as needed
        }
        print(f"User to be deleted: {user_details}")
        # logger.debug(f"User to be deleted: {user_details}")

        # Delete the user
        user.delete()
        return Response(
            {"msg": "User deleted successfully."}, status=status.HTTP_204_NO_CONTENT
        )


#############################
######
######
######
######
######15/07/24 -->EDIT USER FUNCTION
# from rest_framework.views import APIView
# from rest_framework.response import Response
from rest_framework import status
from django.shortcuts import get_object_or_404

# from authentication.models import CustomUser
from .serializers import CustomUserSerializer

# from rest_framework.permissions import IsAdminUser


class EditUserView(APIView):
    permission_classes = [IsAdminUser]

    def patch(self, request, user_id):
        user = get_object_or_404(CustomUser, id=user_id)
        serializer = CustomUserSerializer(user, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def put(self, request, user_id):
        user = get_object_or_404(CustomUser, id=user_id)
        serializer = CustomUserSerializer(user, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


#############################
from django.db.models.fields import CharField, TextField


class PrintUsers(APIView):
    permission_classes = [
        IsAdminOrStaff,
    ]

    def get(self, request):
        role = request.query_params.get("role")
        gender = request.query_params.get("gender")
        search = request.query_params.get("search")

        custom_users = CustomUser.objects.filter(company=request.user.company)
        if role:
            custom_users = custom_users.filter(user_role=role)
        if gender:
            custom_users = custom_users.filter(gender=gender)

        if search:
            search_query = Q()
            for field in CustomUser._meta.get_fields():
                if isinstance(field, (CharField, TextField)):
                    search_query |= Q(**{f"{field.name}__icontains": search})

            custom_users = custom_users.filter(search_query)

        staff_members = Staff.objects.filter(user__company=request.user.company)
        clients = Client.objects.filter(user__company=request.user.company)
        admins = Admin.objects.filter(user__company=request.user.company)

        # Create serializers for each type of user
        custom_user_serializer = CustomUserSerializer(custom_users, many=True)
        staff_serializer = StaffSerializer(staff_members, many=True)
        client_serializer = ClientSerializer(clients, many=True)
        admin_serializer = AdminSerializer(admins, many=True)
        # print(client_serializer.data[0])
        # Combine the serialized data from all user types
        all_users_data = {
            "custom_users": custom_user_serializer.data,
            "staff_members": staff_serializer.data,  # comment at 21/2/2024
            "clients": client_serializer.data,  # comment at 21/2/2024
            "admins": admin_serializer.data,  # comment at 21/2/2024
        }

        return Response(all_users_data, status=status.HTTP_200_OK)


class SearchUsers(generics.ListAPIView):
    permission_classes = [
        IsAdminOrStaff,
    ]
    queryset = CustomUser.objects.all()
    serializer_class = CustomUserSerializer
    filter_backends = [UserSearchFilter]

    def get_queryset(self):
        queryset = super().get_queryset()
        queryset = queryset.filter(company=self.request.user.company)
        return queryset

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        if not request.query_params:
            return Response(
                {"Error": "Please provide data to search."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if not queryset.count():
            return Response(
                {"message": "No users found"}, status=status.HTTP_404_NOT_FOUND
            )
        else:
            serializer = self.get_serializer(queryset, many=True)
            return Response(serializer.data, status=status.HTTP_200_OK)


class PrintAppointments(APIView):
    permission_classes = [
        IsAdminOrStaff,
    ]
    serializer_class = AppointmentSerializer

    def get(self, request):
        # Get the current date
        now = timezone.now().date()
        earlier_date = self.request.query_params.get("earlier_date", None)
        staff = self.request.query_params.get("staff", None)

        # Calculate the date for two months ago
        two_months_ago = now - timedelta(days=60)
        appointments = (
            Appointment.objects.prefetch_related("client", "staff")
            .filter(
                appointment_date__gte=two_months_ago,
                staff__user__company=request.user.company,
            )
            .order_by("-id")
        )
        if earlier_date and staff:
            appointments = appointments.filter(
                earlier_date=earlier_date, staff__id=staff
            )
        serializer = self.serializer_class(appointments, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)


class MyAppointments(APIView):
    # permission_classes = [IsAdminOrStaff,]
    serializer_class = AppointmentSerializer

    def get(self, request):
        appointments = Appointment.objects.filter(client__user=request.user).order_by(
            "-id"
        )
        serializer = self.serializer_class(appointments, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)


@csrf_exempt
def download_excelData(request):
    return render(request, "excelData.html")


class ExportAppointmentsToExcel(APIView):
    permission_classes = [
        IsAdminOrStaff,
    ]

    def get(self, request):
        appointments = Appointment.objects.filter(
            staff__user__company=request.user.company
        )
        wb = Workbook()
        ws = wb.active
        headers = ["Appointment ID", "Client Name", "Staff Name", "Date and Time"]
        ws.append(headers)
        data = []

        # Write user data rows
        for appointment in appointments:
            client = appointment.client.user.username if appointment.client else None
            staff = appointment.staff.user.username if appointment.staff else None
            appointment_date = "null"
            if appointment.appointment_date is not None:
                if isinstance(appointment.appointment_date, str):
                    appointment_date = (
                        datetime.strptime(
                            appointment.appointment_date, "%Y-%m-%d %H:%M:%S"
                        )
                    ).strftime("%Y-%m-%d")
                    # print(hellllo)
                else:
                    appointment_date = appointment.appointment_date.strftime("%Y-%m-%d")
                    # print(hellllo22)
            ws.append([appointment.id, client, staff, appointment_date])
            data.append({appointment.id, client, staff, appointment_date})
        print("data", data)
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        wb.close()
        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="facility_data.xlsx"'
        return response


class ExportEmailSMSToExcel(APIView):
    permission_classes = [IsAdminUser]

    def get(self, request):
        appointments = AppointmentSendInfo.objects.all()
        wb = Workbook()
        ws = wb.active
        headers = [
            "Staff",
            "Client",
            "Message",
            "Date and Time",
            "SMS",
            "Issue SMS",
            "Email",
            "Issue Email",
        ]
        ws.append(headers)

        for appointment in appointments:
            client = f"{appointment.client.user.first_name} {appointment.client.user.last_name}"
            staff = f"{appointment.staff.user.first_name} {appointment.staff.user.last_name}"
            message = appointment.message
            sms = "YES" if appointment.type == SendType.SMS else "NO"
            email = "YES" if appointment.type == SendType.EMAIL else "NO"
            if appointment.send_at is not None:
                if isinstance(appointment.send_at, str):
                    date_time = datetime.strptime(
                        appointment.send_at, "%Y-%m-%d %H:%M:%S"
                    )
                else:
                    date_time = appointment.send_at.strftime("%Y-%m-%d %H:%M:%S")

            ws.append([staff, client, message, date_time, sms, "", email, ""])

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        wb.close()
        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="appointment_data.xlsx"'
        return response


# from io import BytesIO
# # from openpyxl import Workbook
# # from django.http import HttpResponse
# from rest_framework.views import APIView
# from rest_framework.permissions import IsAdminUser
# from .models import Appointment
# from datetime import datetime
# import logging

# logger = logging.getLogger(__name__)

# class ExportAppointmentsToExcel(APIView):
#     permission_classes = [IsAdminUser,]

#     def get(self, request):
#         try:
#             appointments = Appointment.objects.filter(staff__user__company=request.user.company)
#             wb = Workbook()
#             ws = wb.active
#             headers = ['Appointment ID', 'Client Name', 'Staff Name', 'Date', 'Time Start', 'Time End']
#             ws.append(headers)

#             for appointment in appointments:
#                 client = appointment.client.user.username if appointment.client else None
#                 staff = appointment.staff.user.username if appointment.staff else None
#                 appointment_date = "null"
#                 time_start = "null"
#                 time_end = "null"

#                 if appointment.appointment_date is not None:
#                     if isinstance(appointment.appointment_date, str):
#                         appointment_date = datetime.strptime(appointment.appointment_date, '%Y-%m-%d %H:%M:%S').strftime('%Y-%m-%d')
#                         print(test1)
#                     else:
#                         print(test2)
#                         appointment_date = appointment.appointment_date.strftime('%Y-%m-%d')

#                 if hasattr(appointment, 'appointment_start_time_str') and appointment.appointment_start_time_str:
#                     time_start = appointment.appointment_start_time_str

#                 if hasattr(appointment, 'appointment_finish_time_str') and appointment.appointment_finish_time_str:
#                     time_end = appointment.appointment_finish_time_str

#                 ws.append([appointment.id, client, staff, appointment_date, time_start, time_end])

#             buffer = BytesIO()
#             wb.save(buffer)
#             buffer.seek(0)
#             wb.close()

#             response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#             response['Content-Disposition'] = 'attachment; filename="appointments_data.xlsx"'
#             return response

#         except Exception as e:
#             logger.error(f"Error exporting appointments to Excel: {e}")
#             return HttpResponse(status=500, content="Internal Server Error")
from django.core.mail import EmailMessage
from django.template.loader import render_to_string
from django.utils.html import strip_tags


class SendEmailView(APIView):
    permission_classes = [IsAdminOrStaff]

    def post(self, request):
        subject = "Appointment Created"
        template = "email/appointment_created_email.html"
        appointment_id = request.data.get("appointment_id")

        if not (subject and template):
            return Response(
                {
                    "error": "All fields are required: subject, template, appointment_id, recipient_list"
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            appointment = Appointment.objects.get(pk=appointment_id)
        except Appointment.DoesNotExist:
            return Response(
                {"error": "Appointment not found"}, status=status.HTTP_404_NOT_FOUND
            )

        context = {"appointment": appointment}
        message = render_to_string(template, context)
        message_text = strip_tags(message)

        error_message = None
        try:
            email = EmailMessage(subject, message, to=[appointment.client.user.email])
            email.content_subtype = "html"  # Specify the content type as HTML
            email.send()
        except Exception as e:
            error_message = str(e)

        AppointmentSendInfo.objects.create(
            staff=appointment.staff,
            client=appointment.client,
            message=message_text,
            type=SendType.EMAIL,
            send_at=timezone.now(),
            error=error_message,
        )

        if error_message:
            return Response(
                {"error": "Failed to send email", "details": error_message},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        return Response("Email sent successfully", status=status.HTTP_200_OK)


class AppointmentEditView(APIView):
    permission_classes = [IsAdminOrStaff]
    serializer_class = AppointmentEditSerializer

    def put(self, request, pk):
        return self.update_appointment(request, pk, partial=False)

    def patch(self, request, pk):
        return self.update_appointment(request, pk, partial=True)

    def update_appointment(self, request, pk, partial):
        try:
            appointment = Appointment.objects.get(pk=pk)
        except Appointment.DoesNotExist:
            return Response(
                {"error": "Appointment not found"}, status=status.HTTP_404_NOT_FOUND
            )

        appointment_serializer = self.serializer_class(
            appointment,
            data=request.data,
            context={"user": request.user},
            partial=partial,
        )
        if appointment_serializer.is_valid():
            appointment_serializer.save()
            return Response(
                "Appointment updated successfully", status=status.HTTP_200_OK
            )
        else:
            return Response(
                appointment_serializer.errors, status=status.HTTP_400_BAD_REQUEST
            )


class ClientAppointmentView(APIView):
    serializer_class = AppointmentSerializer
    permission_classes = [IsAuthenticated]

    def get(self, request, pk):
        try:
            client = Client.objects.get(user__pk=pk)
            appointments = Appointment.objects.filter(client=client)
            serializer = self.serializer_class(appointments, many=True)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except Appointment.DoesNotExist:
            return Response(
                {"error": "Appointment not found"}, status=status.HTTP_404_NOT_FOUND
            )


class AppointmentView(APIView):
    permission_classes = [
        IsAdminOrStaff,
    ]
    serializer_class = AppointmentSerializer

    def send_email(self, subject, template, context, recipient_list):
        message = render_to_string(template, context)
        email = EmailMessage(subject, message, to=recipient_list)
        email.content_subtype = "html"  # Specify the content type as HTML
        email.send()

    def get(self, request, pk):
        try:
            appointment = Appointment.objects.get(pk=pk)
            serializer = self.serializer_class(appointment)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except Appointment.DoesNotExist:
            return Response(
                {"error": "Appointment not found"}, status=status.HTTP_404_NOT_FOUND
            )

    def post(self, request):
        appointment_serializer = self.serializer_class(
            data=request.data, context={"user": request.user}
        )
        if appointment_serializer.is_valid():
            appointment_serializer.save()

            return Response(appointment_serializer.data, status=status.HTTP_201_CREATED)
        else:
            return Response(
                appointment_serializer.errors, status=status.HTTP_400_BAD_REQUEST
            )

    def put(self, request, pk):
        try:
            appointment = Appointment.objects.get(pk=pk)
        except Appointment.DoesNotExist:
            return Response(
                {"error": "Appointment not found"}, status=status.HTTP_404_NOT_FOUND
            )

        appointment_serializer = self.serializer_class(
            appointment, data=request.data, context={"user": request.user}, partial=True
        )
        if appointment_serializer.is_valid():
            appointment_serializer.save()
            return Response(
                "Appointment updated successfully", status=status.HTTP_200_OK
            )
        else:
            return Response(
                appointment_serializer.errors, status=status.HTTP_400_BAD_REQUEST
            )

    def delete(self, request, pk):
        try:
            appointment = Appointment.objects.get(pk=pk)
        except Appointment.DoesNotExist:
            return Response(
                {"error": "Appointment not found"}, status=status.HTTP_404_NOT_FOUND
            )

        appointment.delete()
        return Response(
            "Appointment deleted successfully", status=status.HTTP_204_NO_CONTENT
        )


###############################################################################15/07/24
class EditAppointment(APIView):
    print("3")
    permission_classes = [
        IsAdminOrStaff,
    ]
    serializer_class = AppointmentSerializer
    print("1")

    def put(self, request, appointment_id):
        appointment = get_object_or_404(Appointment, id=appointment_id)
        appointment_serializer = self.serializer_class(
            appointment, data=request.data, context={"user": request.user}
        )
        if appointment_serializer.is_valid():
            appointment_serializer.save()
            return Response(
                "Appointment updated successfully", status=status.HTTP_200_OK
            )
        else:
            return Response(
                appointment_serializer.errors, status=status.HTTP_400_BAD_REQUEST
            )

    print("2")

    def patch(self, request, appointment_id):
        appointment = get_object_or_404(Appointment, id=appointment_id)
        appointment_serializer = self.serializer_class(
            appointment, data=request.data, partial=True, context={"user": request.user}
        )
        if appointment_serializer.is_valid():
            appointment_serializer.save()
            return Response(
                "Appointment partially updated successfully", status=status.HTTP_200_OK
            )
        else:
            return Response(
                appointment_serializer.errors, status=status.HTTP_400_BAD_REQUEST
            )


class CreateAppointmentCommentView(APIView):
    permission_classes = [
        IsAuthenticated
    ]  # Only authenticated users can create comments

    def post(self, request, *args, **kwargs):
        print(request.data)
        serializer = AppointmentCommentSerializer(data=request.data)
        if serializer.is_valid():
            print(request.user)
            serializer.save(
                user=request.user
            )  # Automatically set the user to the current authenticated user
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class AppointmentCommentListView(generics.ListAPIView):
    def get_queryset(self):
        appointment_id = self.kwargs.get("appointment_id")
        if appointment_id:
            # Return the queryset of comments for the specific appointment
            return AppointmentComment.objects.filter(appointment__id=appointment_id)
        return AppointmentComment.objects.none()

    def list(self, request, *args, **kwargs):
        queryset = self.get_queryset()

        comments = []
        for comment in queryset:
            comments.append(
                {
                    "appointment": {
                        "id": comment.appointment.id,
                        "date": comment.appointment.appointment_date,
                        "start_time": comment.appointment.appointment_start,
                        "end_time": comment.appointment.appointment_end,
                    },
                    "user": {
                        "id": comment.user.id,
                        "username": comment.user.username,
                        "email": comment.user.email,
                    },
                    "id": comment.id,
                    "visibility": comment.visibility,
                    "comment_text": comment.comment_text,
                    "created_at": comment.created_at,
                }
            )
        print(comments)
        return Response(comments, status=status.HTTP_200_OK)


# class DeleteAppointment(APIView):
#     permission_classes = [IsAdminOrStaff,]

#     def delete(self, request, appointment_id):
#         appointment = get_object_or_404(Appointment, id=appointment_id)
#         appointment.delete()
#         return Response("Appointment deleted successfully", status=status.HTTP_204_NO_CONTENT)
#########################################################################
# from rest_framework.views import APIView
# from rest_framework.response import Response
# from rest_framework import status
# from django.shortcuts import get_object_or_404
# from .models import Appointment  # Assuming you have an Appointment model
# from .permissions import IsAdminOrStaff  # Assuming you have this custom permission

# class DeleteAppointment(APIView):
#     permission_classes = [IsAdminOrStaff,]  # Replace with your custom permission class if needed

#     def delete(self, request, appointment_id):
#         # Fetch the appointment object to be deleted
#         appointment = get_object_or_404(Appointment, id=appointment_id)

#         # Perform the deletion
#         appointment.delete()

#         # Return a response indicating success
#         return Response("Appointment deleted successfully", status=status.HTTP_204_NO_CONTENT)

###############################################################################
# class AssignClientToStaff(APIView):
#     permission_classes = [IsAdminOrStaff,]
#     serializer_class = AssignmentSerializer

#     user_serializer = CustomUserSerializer


#     def post(self,request):
#         client_id = request.data.get("client_id")
#         staff_id = request.data.get("staff_id")
#         user_serializer = self.serializer_class(data = request.data,context = { 'user': request.user})

#         user = get_object_or_404(Appointment, id=appointment_id, )
#         appointment_serializer = self.serializer_class(appointment, data=request.data, context={'user': request.user})

#         if (user_serializer.is_valid):
#             user_serializer.save()

#             return Response(f'Client with id {client_id} assigned to staff with id {staff_id}', status = status.HTTP_200_OK)
#         else:
#             print("costis")
#             return Response(appointment_serializer.errors,status = status.HTTP_400_BAD_REQUEST)
############################################################################################################
# class AssignClientToStaff(APIView):
#     permission_classes = [IsAdminOrStaff]
#     serializer_class = AssignmentSerializer

#     def post(self, request):
#         client_id = request.data.get("client_id")
#         staff_id = request.data.get("staff_id")
#         assignment_serializer = self.serializer_class(data=request.data, context={'user': request.user})

#         if assignment_serializer.is_valid():
#             assignment_serializer.save()
#             return Response(f'Client with id {client_id} assigned to staff with id {staff_id}', status=status.HTTP_200_OK)
#         else:
#             return Response(assignment_serializer.errors, status=status.HTTP_400_BAD_REQUEST)
#####################################################################################################


class AssignClientToStaff(APIView):
    permission_classes = [IsAdminOrStaff]
    serializer_class = AssignmentSerializer

    def post(self, request):
        client_id = request.data.get("client_id")
        staff_id = request.data.get("staff_id")

        try:
            client = get_object_or_404(Client, pk=client_id)
            staff = get_object_or_404(Staff, pk=staff_id)
            client.assigned_staff = staff
            client.save()
            return Response(status=status.HTTP_200_OK)
        except:
            return Response(status=status.HTTP_400_BAD_REQUEST)


class GetStaffWithAssignedClients(APIView):
    permission_classes = [IsAdminOrStaff]

    def get(self, request):
        staff_members = Staff.objects.prefetch_related("client_set__user").filter(
            user__company=request.user.company
        )
        staff_clients_data = []
        for staff in staff_members:
            assigned_clients = staff.client_set.all()
            if assigned_clients:
                clients_info = [
                    {"client_user_name": client.user.username}
                    for client in assigned_clients
                ]
                staff_clients_data.append(
                    {
                        "staff_user_name": staff.user.username,
                        "assigned_clients": clients_info,
                    }
                )
        return Response({"staff_members": staff_clients_data}, status=200)


#################################################################13/7/24
import logging
from rest_framework.views import APIView
from rest_framework.response import Response

# from rest_framework import status
from rest_framework.pagination import PageNumberPagination

# Set up logging
logger = logging.getLogger(__name__)


class GetCombinedUserAndStaffClients(APIView):
    permission_classes = [
        IsAdminOrStaff,
    ]

    def get(self, request):
        logger.info("GetCombinedUserAndStaffClients view called")
        try:
            paginator = PageNumberPagination()
            paginator.page_size = 10

            # Get all custom users, staff members, clients, and admins
            custom_users = CustomUser.objects.filter(company=request.user.company)
            staff_members = Staff.objects.filter(user__company=request.user.company)
            clients = Client.objects.filter(user__company=request.user.company)
            admins = Admin.objects.filter(user__company=request.user.company)

            # Serialize each type of user
            custom_user_serializer = CustomUserSerializer(custom_users, many=True)
            staff_serializer = StaffSerializer(staff_members, many=True)
            client_serializer = ClientSerializer(clients, many=True)
            admin_serializer = AdminSerializer(admins, many=True)

            # Paginate serialized data
            custom_users_paginated = paginator.paginate_queryset(
                custom_user_serializer.data, request
            )
            staff_members_paginated = paginator.paginate_queryset(
                staff_serializer.data, request
            )
            clients_paginated = paginator.paginate_queryset(
                client_serializer.data, request
            )
            admins_paginated = paginator.paginate_queryset(
                admin_serializer.data, request
            )

            # Get staff members with their assigned clients
            staff_with_clients = Staff.objects.prefetch_related(
                "client_set__user"
            ).filter(user__company=request.user.company)
            staff_clients_data = []
            for staff in staff_with_clients:
                assigned_clients = staff.client_set.all()
                if assigned_clients:
                    clients_info = [
                        {"client_user_name": client.user.username}
                        for client in assigned_clients
                    ]
                    staff_clients_data.append(
                        {
                            "staff_user_name": staff.user.username,
                            "assigned_clients": clients_info,
                        }
                    )

            # Paginate staff with assigned clients data if necessary
            staff_with_clients_paginated = paginator.paginate_queryset(
                staff_clients_data, request
            )

            # Combine the serialized data from all user types and staff with assigned clients
            all_users_data = {
                "custom_users": custom_users_paginated,
                "staff_members": staff_members_paginated,
                "clients": clients_paginated,
                "admins": admins_paginated,
                "staff_with_assigned_clients": staff_with_clients_paginated,
            }

            logger.info("Data successfully retrieved and serialized")
            return paginator.get_paginated_response(all_users_data)

        except Exception as e:
            logger.error(f"Error in GetCombinedUserAndStaffClients view: {e}")
            return Response(
                {"error": "An error occurred"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


############################################################################


class SendAppointmentInfo(APIView):
    permission_classes = [IsAdminOrStaff]

    def post(self, request, appointment_id):
        if Appointment.objects.filter(pk=appointment_id).exists():
            appointment = Appointment.objects.get(pk=appointment_id)
            user_company = request.user.company.pk if request.user.company else None
            staff_company = (
                appointment.staff.user.company.pk
                if appointment.staff.user.company
                else None
            )
            if staff_company != user_company:
                return Response(
                    {
                        "message": "Action Reserved only for administrators or staff of the same company"
                    },
                    status=status.HTTP_403_FORBIDDEN,
                )
            client = appointment.client
            staff = appointment.staff
            if not client.user.phone_number:
                return Response(
                    "Client does not have a number", status=status.HTTP_400_BAD_REQUEST
                )

            phone = str(client.user.phone_number).replace("+", "")
            sender = "Appt Info"
            key = settings.SMS_API_KEY
            message = (
                f"Hi, {client.user.first_name} {client.user.last_name}. "
                f"You have an appointment with {staff.user.first_name} {staff.user.last_name} "
                f"at {appointment.appointment_start.strftime('%H:%M')} on {appointment.appointment_date.strftime('%d-%m-%Y')}."
            )
            url = f"https://websms.com.cy/api/send-sm?to={phone}&from={sender}&key={key}&encoding=GSM&message={message}"

            error_message = None
            try:
                response = requests.get(url)
                response_data = response.json()
                if response_data.get("error"):
                    error_message = response_data.get("error")
                    return Response(
                        response_data.get("error"), status=status.HTTP_400_BAD_REQUEST
                    )
            except Exception as e:
                error_message = str(e)
                return Response(
                    f"Failed to send SMS: {error_message}",
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR,
                )

            AppointmentSendInfo.objects.create(
                staff=appointment.staff,
                client=appointment.client,
                message=message,
                type=SendType.SMS,
                send_at=timezone.now(),
                error=error_message,
            )
            return Response(
                "Appointment reminder sent successfully", status=status.HTTP_200_OK
            )
        else:
            return Response("Appointment not found", status=status.HTTP_404_NOT_FOUND)


class GeneralClientStats(APIView):
    permission_classes = [
        IsAdminOrStaff,
    ]

    def get(self, request):
        today = now().date()
        start_of_year = datetime(today.year, 1, 1).date()
        end_of_year = datetime(today.year, 12, 31).date()
        start_of_month = datetime(today.year, today.month, 1).date()
        end_of_month = (
            (datetime(today.year, today.month + 1, 1) - timedelta(days=1)).date()
            if today.month < 12
            else datetime(today.year, 12, 31).date()
        )
        start_of_week = today - timedelta(days=today.weekday())
        end_of_week = start_of_week + timedelta(days=6)
        data = {
            "all": {
                "week": Client.objects.filter(
                    user__company=request.user.company,
                    user__date_joined__date__range=[start_of_week, end_of_week],
                ).count(),
                "month": Client.objects.filter(
                    user__company=request.user.company,
                    user__date_joined__date__range=[start_of_month, end_of_month],
                ).count(),
                "year": Client.objects.filter(
                    user__company=request.user.company,
                    user__date_joined__date__range=[start_of_year, end_of_year],
                ).count(),
            },
            "male": {
                "week": Client.objects.filter(
                    user__company=request.user.company,
                    user__date_joined__date__range=[start_of_week, end_of_week],
                    user__gender="M",
                ).count(),
                "month": Client.objects.filter(
                    user__company=request.user.company,
                    user__date_joined__date__range=[start_of_month, end_of_month],
                    user__gender="M",
                ).count(),
                "year": Client.objects.filter(
                    user__company=request.user.company,
                    user__date_joined__date__range=[start_of_year, end_of_year],
                    user__gender="M",
                ).count(),
            },
            "female": {
                "week": Client.objects.filter(
                    user__company=request.user.company,
                    user__date_joined__date__range=[start_of_week, end_of_week],
                    user__gender="F",
                ).count(),
                "month": Client.objects.filter(
                    user__company=request.user.company,
                    user__date_joined__date__range=[start_of_month, end_of_month],
                    user__gender="F",
                ).count(),
                "year": Client.objects.filter(
                    user__company=request.user.company,
                    user__date_joined__date__range=[start_of_year, end_of_year],
                    user__gender="F",
                ).count(),
            },
        }
        return Response(data, status=status.HTTP_200_OK)


class ClientStatsByWeek(APIView):
    permission_classes = [
        IsAdminOrStaff,
    ]

    def get(self, request, weekstart):
        format_string = "%Y-%m-%d"
        try:
            is_valid = datetime.strptime(weekstart, format_string)
            if not is_valid:
                return Response(
                    {"weekstart": "Invalid date"}, status=status.HTTP_400_BAD_REQUEST
                )
        except Exception as e:
            print("Exception = ", str(e))
            return Response(
                {"weekstart": "Invalid date"}, status=status.HTTP_400_BAD_REQUEST
            )
        start_of_week = datetime.strptime(weekstart, format_string)
        if not start_of_week.weekday() == 0:
            return Response(
                {"weekstart": "date is not start of the week (Monday)"},
                status=status.HTTP_400_BAD_REQUEST,
            )
        end_of_week = start_of_week + timedelta(days=6)
        data = {
            "all": Client.objects.filter(
                user__company=request.user.company,
                user__date_joined__date__range=[start_of_week, end_of_week],
            ).count(),
            "male": Client.objects.filter(
                user__company=request.user.company,
                user__date_joined__date__range=[start_of_week, end_of_week],
                user__gender="M",
            ).count(),
            "female": Client.objects.filter(
                user__company=request.user.company,
                user__date_joined__date__range=[start_of_week, end_of_week],
                user__gender="F",
            ).count(),
        }
        return Response(data, status=status.HTTP_200_OK)


def is_valid_date(date_string):
    if not bool(re.match(r"^\d{2}-\d{4}$", date_string)):
        return False

    month, year = date_string.split("-")
    try:
        datetime(
            year=int(year), month=int(month), day=1
        )  # Using day=1 since only month and year matter
        return True
    except Exception as e:
        return False


class ClientStatsByMonth(APIView):
    permission_classes = [
        IsAdminOrStaff,
    ]

    def get(self, request, month_string):
        if not is_valid_date(month_string):
            return Response(
                {"month": "Invalid date"}, status=status.HTTP_400_BAD_REQUEST
            )
        month, year = month_string.split("-")
        month = int(month)
        year = int(year)
        start_of_month = datetime(year, month, 1).date()
        end_of_month = (
            (datetime(year, month + 1, 1) - timedelta(days=1)).date()
            if month < 12
            else datetime(year, 12, 31).date()
        )
        data = {
            "all": Client.objects.filter(
                user__company=request.user.company,
                user__date_joined__date__range=[start_of_month, end_of_month],
            ).count(),
            "male": Client.objects.filter(
                user__company=request.user.company,
                user__date_joined__date__range=[start_of_month, end_of_month],
                user__gender="M",
            ).count(),
            "female": Client.objects.filter(
                user__company=request.user.company,
                user__date_joined__date__range=[start_of_month, end_of_month],
                user__gender="F",
            ).count(),
        }
        return Response(data, status=status.HTTP_200_OK)


class ClientStatsByYear(APIView):
    permission_classes = [
        IsAdminOrStaff,
    ]

    def get(self, request, year):
        if not re.match(r"^\d{4}$", str(year)):
            return Response(
                {"year": "Invalid date"}, status=status.HTTP_400_BAD_REQUEST
            )
        start_of_year = datetime(year, 1, 1).date()
        end_of_year = datetime(year, 12, 31).date()
        data = {
            "all": Client.objects.filter(
                user__company=request.user.company,
                user__date_joined__date__range=[start_of_year, end_of_year],
            ).count(),
            "male": Client.objects.filter(
                user__company=request.user.company,
                user__date_joined__date__range=[start_of_year, end_of_year],
                user__gender="M",
            ).count(),
            "female": Client.objects.filter(
                user__company=request.user.company,
                user__date_joined__date__range=[start_of_year, end_of_year],
                user__gender="F",
            ).count(),
        }
        return Response(data, status=status.HTTP_200_OK)


class GeneralAppointmentStats(APIView):
    permission_classes = [
        IsAdminOrStaff,
    ]

    def get(self, request):
        today = now().date()
        start_of_year = datetime(today.year, 1, 1).date()
        end_of_year = datetime(today.year, 12, 31).date()
        start_of_month = datetime(today.year, today.month, 1).date()
        end_of_month = (
            (datetime(today.year, today.month + 1, 1) - timedelta(days=1)).date()
            if today.month < 12
            else datetime(today.year, 12, 31).date()
        )
        start_of_week = today - timedelta(days=today.weekday())
        end_of_week = start_of_week + timedelta(days=6)
        data = {
            "week": [
                {
                    "id": staff.id,
                    "total": Appointment.objects.filter(
                        staff=staff.id,
                        appointment_date__range=[start_of_week, end_of_week],
                    ).count(),
                }
                for staff in Staff.objects.filter(user__company=request.user.company)
            ],
            "month": [
                {
                    "id": staff.id,
                    "total": Appointment.objects.filter(
                        staff=staff.id,
                        appointment_date__range=[start_of_month, end_of_month],
                    ).count(),
                }
                for staff in Staff.objects.filter(user__company=request.user.company)
            ],
            "year": [
                {
                    "id": staff.id,
                    "total": Appointment.objects.filter(
                        staff=staff.id,
                        appointment_date__range=[start_of_year, end_of_year],
                    ).count(),
                }
                for staff in Staff.objects.filter(user__company=request.user.company)
            ],
        }
        return Response(data, status=status.HTTP_200_OK)


class AppointmentStatsByWeek(APIView):
    permission_classes = [
        IsAdminOrStaff,
    ]

    def get(self, request, weekstart):
        format_string = "%Y-%m-%d"
        try:
            is_valid = datetime.strptime(weekstart, format_string)
            if not is_valid:
                return Response(
                    {"weekstart": "Invalid date"}, status=status.HTTP_400_BAD_REQUEST
                )
        except:
            return Response(
                {"weekstart": "Invalid date"}, status=status.HTTP_400_BAD_REQUEST
            )
        start_of_week = datetime.strptime(weekstart, format_string)
        if not start_of_week.weekday() == 0:
            return Response(
                {"weekstart": "date is not start of the week (Monday)"},
                status=status.HTTP_400_BAD_REQUEST,
            )
        end_of_week = start_of_week + timedelta(days=6)
        data = [
            {
                "id": staff.id,
                "total": Appointment.objects.filter(
                    staff=staff.id, appointment_date__range=[start_of_week, end_of_week]
                ).count(),
            }
            for staff in Staff.objects.filter(user__company=request.user.company)
        ]
        return Response(data, status=status.HTTP_200_OK)


class AppointmentStatsByMonth(APIView):
    permission_classes = [
        IsAdminOrStaff,
    ]

    def get(self, request, month_string):
        if not is_valid_date(month_string):
            return Response(
                {"month": "Invalid date"}, status=status.HTTP_400_BAD_REQUEST
            )
        month, year = month_string.split("-")
        month = int(month)
        year = int(year)
        start_of_month = datetime(year, month, 1).date()
        end_of_month = (
            (datetime(year, month + 1, 1) - timedelta(days=1)).date()
            if month < 12
            else datetime(year, 12, 31).date()
        )
        data = [
            {
                "id": staff.id,
                "total": Appointment.objects.filter(
                    staff=staff.id,
                    appointment_date__range=[start_of_month, end_of_month],
                ).count(),
            }
            for staff in Staff.objects.filter(user__company=request.user.company)
        ]
        return Response(data, status=status.HTTP_200_OK)


class AppointmentStatsByYear(APIView):
    permission_classes = [
        IsAdminOrStaff,
    ]

    def get(self, request, year):
        if not re.match(r"^\d{4}$", str(year)):
            return Response(
                {"year": "Invalid date"}, status=status.HTTP_400_BAD_REQUEST
            )
        start_of_year = datetime(year, 1, 1).date()
        end_of_year = datetime(year, 12, 31).date()
        data = [
            {
                "id": staff.id,
                "total": Appointment.objects.filter(
                    staff=staff.id, appointment_date__range=[start_of_year, end_of_year]
                ).count(),
            }
            for staff in Staff.objects.filter(user__company=request.user.company)
        ]
        return Response(data, status=status.HTTP_200_OK)


class GeneralAssignedClientStats(APIView):
    permission_classes = [
        IsAdminOrStaff,
    ]

    def get(self, request):
        today = now().date()
        start_of_year = datetime(today.year, 1, 1).date()
        end_of_year = datetime(today.year, 12, 31).date()
        start_of_month = datetime(today.year, today.month, 1).date()
        end_of_month = (
            (datetime(today.year, today.month + 1, 1) - timedelta(days=1)).date()
            if today.month < 12
            else datetime(today.year, 12, 31).date()
        )
        start_of_week = today - timedelta(days=today.weekday())
        end_of_week = start_of_week + timedelta(days=6)
        data = {
            "week": [
                {
                    "id": staff.id,
                    "total": Client.objects.filter(
                        assigned_staff=staff.id,
                        user__date_joined__date__range=[start_of_week, end_of_week],
                    ).count(),
                }
                for staff in Staff.objects.filter(user__company=request.user.company)
            ],
            "month": [
                {
                    "id": staff.id,
                    "total": Client.objects.filter(
                        assigned_staff=staff.id,
                        user__date_joined__date__range=[start_of_month, end_of_month],
                    ).count(),
                }
                for staff in Staff.objects.filter(user__company=request.user.company)
            ],
            "year": [
                {
                    "id": staff.id,
                    "total": Client.objects.filter(
                        assigned_staff=staff.id,
                        user__date_joined__date__range=[start_of_year, end_of_year],
                    ).count(),
                }
                for staff in Staff.objects.filter(user__company=request.user.company)
            ],
        }
        return Response(data, status=status.HTTP_200_OK)


class AssignedClientStatsByWeek(APIView):
    permission_classes = [
        IsAdminOrStaff,
    ]

    def get(self, request, weekstart):
        format_string = "%Y-%m-%d"
        try:
            is_valid = datetime.strptime(weekstart, format_string)
            if not is_valid:
                return Response(
                    {"weekstart": "Invalid date"}, status=status.HTTP_400_BAD_REQUEST
                )
        except:
            return Response(
                {"weekstart": "Invalid date"}, status=status.HTTP_400_BAD_REQUEST
            )
        start_of_week = datetime.strptime(weekstart, format_string)
        if not start_of_week.weekday() == 0:
            return Response(
                {"weekstart": "date is not start of the week (Monday)"},
                status=status.HTTP_400_BAD_REQUEST,
            )
        end_of_week = start_of_week + timedelta(days=6)
        data = [
            {
                "id": staff.id,
                "total": Client.objects.filter(
                    assigned_staff=staff.id,
                    user__date_joined__date__range=[start_of_week, end_of_week],
                ).count(),
            }
            for staff in Staff.objects.filter(user__company=request.user.company)
        ]
        return Response(data, status=status.HTTP_200_OK)


class AssignedClientStatsByMonth(APIView):
    permission_classes = [
        IsAdminOrStaff,
    ]

    def get(self, request, month_string):
        if not is_valid_date(month_string):
            return Response(
                {"month": "Invalid date"}, status=status.HTTP_400_BAD_REQUEST
            )
        month, year = month_string.split("-")
        month = int(month)
        year = int(year)
        start_of_month = datetime(year, month, 1).date()
        end_of_month = (
            (datetime(year, month + 1, 1) - timedelta(days=1)).date()
            if month < 12
            else datetime(year, 12, 31).date()
        )
        data = [
            {
                "id": staff.id,
                "total": Client.objects.filter(
                    assigned_staff=staff.id,
                    user__date_joined__date__range=[start_of_month, end_of_month],
                ).count(),
            }
            for staff in Staff.objects.filter(user__company=request.user.company)
        ]
        return Response(data, status=status.HTTP_200_OK)


class AssignedClientStatsByYear(APIView):
    permission_classes = [
        IsAdminOrStaff,
    ]

    def get(self, request, year):
        start_of_year = datetime(year, 1, 1).date()
        end_of_year = datetime(year, 12, 31).date()
        data = [
            {
                "id": staff.id,
                "total": Client.objects.filter(
                    assigned_staff=staff.id,
                    user__date_joined__date__range=[start_of_year, end_of_year],
                ).count(),
            }
            for staff in Staff.objects.filter(user__company=request.user.company)
        ]
        return Response(data, status=status.HTTP_200_OK)


class StaffUserListView(ListAPIView):
    serializer_class = StaffUserSerializer
    queryset = Staff.objects.all()
    permission_classes = [IsAdminOrStaff]


class ClientUserListView(ListAPIView):
    serializer_class = ClientUserSerializer
    queryset = Client.objects.all()
    permission_classes = [IsAdminOrStaff]


###################################


class GeneralClientStats(APIView):
    permission_classes = [
        IsAdminOrStaff,
    ]

    def get(self, request):
        today = now().date()
        start_of_year = datetime(today.year, 1, 1).date()
        end_of_year = datetime(today.year, 12, 31).date()
        start_of_month = datetime(today.year, today.month, 1).date()
        end_of_month = (
            (datetime(today.year, today.month + 1, 1) - timedelta(days=1)).date()
            if today.month < 12
            else datetime(today.year, 12, 31).date()
        )
        start_of_week = today - timedelta(days=today.weekday())
        end_of_week = start_of_week + timedelta(days=6)
        data = {
            "all": {
                "week": Client.objects.filter(
                    user__company=request.user.company,
                    user__date_joined__date__range=[start_of_week, end_of_week],
                ).count(),
                "month": Client.objects.filter(
                    user__company=request.user.company,
                    user__date_joined__date__range=[start_of_month, end_of_month],
                ).count(),
                "year": Client.objects.filter(
                    user__company=request.user.company,
                    user__date_joined__date__range=[start_of_year, end_of_year],
                ).count(),
            },
            "male": {
                "week": Client.objects.filter(
                    user__company=request.user.company,
                    user__date_joined__date__range=[start_of_week, end_of_week],
                    user__gender="M",
                ).count(),
                "month": Client.objects.filter(
                    user__company=request.user.company,
                    user__date_joined__date__range=[start_of_month, end_of_month],
                    user__gender="M",
                ).count(),
                "year": Client.objects.filter(
                    user__company=request.user.company,
                    user__date_joined__date__range=[start_of_year, end_of_year],
                    user__gender="M",
                ).count(),
            },
            "female": {
                "week": Client.objects.filter(
                    user__company=request.user.company,
                    user__date_joined__date__range=[start_of_week, end_of_week],
                    user__gender="F",
                ).count(),
                "month": Client.objects.filter(
                    user__company=request.user.company,
                    user__date_joined__date__range=[start_of_month, end_of_month],
                    user__gender="F",
                ).count(),
                "year": Client.objects.filter(
                    user__company=request.user.company,
                    user__date_joined__date__range=[start_of_year, end_of_year],
                    user__gender="F",
                ).count(),
            },
        }
        return Response(data, status=status.HTTP_200_OK)


class ClientStatsByWeek(APIView):
    permission_classes = [
        IsAdminOrStaff,
    ]

    def get(self, request, weekstart):
        format_string = "%Y-%m-%d"
        try:
            is_valid = datetime.strptime(weekstart, format_string)
            if not is_valid:
                return Response(
                    {"weekstart": "Invalid date"}, status=status.HTTP_400_BAD_REQUEST
                )
        except Exception as e:
            print("Exception = ", str(e))
            return Response(
                {"weekstart": "Invalid date"}, status=status.HTTP_400_BAD_REQUEST
            )
        start_of_week = datetime.strptime(weekstart, format_string)
        if not start_of_week.weekday() == 0:
            return Response(
                {"weekstart": "date is not start of the week (Monday)"},
                status=status.HTTP_400_BAD_REQUEST,
            )
        end_of_week = start_of_week + timedelta(days=6)

        print(start_of_week, end_of_week, "===")

        data = {
            "all": Client.objects.filter(
                user__company=request.user.company,
                user__date_joined__date__range=[start_of_week, end_of_week],
            ).count(),
            "male": Client.objects.filter(
                user__company=request.user.company,
                user__date_joined__date__range=[start_of_week, end_of_week],
                user__gender="M",
            ).count(),
            "female": Client.objects.filter(
                user__company=request.user.company,
                user__date_joined__date__range=[start_of_week, end_of_week],
                user__gender="F",
            ).count(),
        }
        return Response(data, status=status.HTTP_200_OK)


def is_valid_date(date_string):
    if not bool(re.match(r"^\d{2}-\d{4}$", date_string)):
        return False

    month, year = date_string.split("-")
    try:
        datetime(
            year=int(year), month=int(month), day=1
        )  # Using day=1 since only month and year matter
        return True
    except Exception as e:
        return False


class ClientStatsByMonth(APIView):
    permission_classes = [
        IsAdminOrStaff,
    ]

    def get(self, request, month_string):
        if not is_valid_date(month_string):
            return Response(
                {"month": "Invalid date"}, status=status.HTTP_400_BAD_REQUEST
            )
        month, year = month_string.split("-")
        month = int(month)
        year = int(year)
        start_of_month = datetime(year, month, 1).date()
        end_of_month = (
            (datetime(year, month + 1, 1) - timedelta(days=1)).date()
            if month < 12
            else datetime(year, 12, 31).date()
        )
        data = {
            "all": Client.objects.filter(
                user__company=request.user.company,
                user__date_joined__date__range=[start_of_month, end_of_month],
            ).count(),
            "male": Client.objects.filter(
                user__company=request.user.company,
                user__date_joined__date__range=[start_of_month, end_of_month],
                user__gender="M",
            ).count(),
            "female": Client.objects.filter(
                user__company=request.user.company,
                user__date_joined__date__range=[start_of_month, end_of_month],
                user__gender="F",
            ).count(),
        }
        return Response(data, status=status.HTTP_200_OK)


class ClientStatsByYear(APIView):
    permission_classes = [
        IsAdminOrStaff,
    ]

    def get(self, request, year):
        if not re.match(r"^\d{4}$", str(year)):
            return Response(
                {"year": "Invalid date"}, status=status.HTTP_400_BAD_REQUEST
            )
        start_of_year = datetime(year, 1, 1).date()
        end_of_year = datetime(year, 12, 31).date()
        data = {
            "all": Client.objects.filter(
                user__company=request.user.company,
                user__date_joined__date__range=[start_of_year, end_of_year],
            ).count(),
            "male": Client.objects.filter(
                user__company=request.user.company,
                user__date_joined__date__range=[start_of_year, end_of_year],
                user__gender="M",
            ).count(),
            "female": Client.objects.filter(
                user__company=request.user.company,
                user__date_joined__date__range=[start_of_year, end_of_year],
                user__gender="F",
            ).count(),
        }
        return Response(data, status=status.HTTP_200_OK)


# class GeneralAppointmentStats(APIView):
#     permission_classes = [IsAdminOrStaff,]

#     def get(self,request):
#         today = now().date()
#         start_of_year = datetime(today.year, 1, 1).date()
#         end_of_year = datetime(today.year, 12, 31).date()
#         start_of_month = datetime(today.year, today.month, 1).date()
#         end_of_month = (datetime(today.year, today.month + 1, 1) - timedelta(days=1)).date() if today.month < 12 else datetime(today.year, 12, 31).date()
#         start_of_week = today - timedelta(days=today.weekday())
#         end_of_week = start_of_week + timedelta(days=6)
#         data = {
#             "week": [{"id": staff.id, "total": Appointment.objects.filter(staff = staff.id, appointment_date__range=[start_of_week, end_of_week]).count()} for staff in  Staff.objects.filter(user__company = request.user.company)],
#             "month": [{"id": staff.id, "total": Appointment.objects.filter(staff = staff.id, appointment_date__range=[start_of_month, end_of_month]).count()} for staff in  Staff.objects.filter(user__company = request.user.company)],
#             "year": [{"id": staff.id, "total": Appointment.objects.filter(staff = staff.id, appointment_date__range=[start_of_year, end_of_year]).count()} for staff in  Staff.objects.filter(user__company = request.user.company)]
#         }
#         return Response(data, status = status.HTTP_200_OK)


class GeneralAppointmentStats(APIView):
    permission_classes = [
        IsAdminOrStaff,
    ]

    def get(self, request):
        today = now().date()

        # Calculate time periods
        start_of_year = datetime(today.year, 1, 1).date()
        end_of_year = datetime(today.year, 12, 31).date()

        start_of_month = datetime(today.year, today.month, 1).date()
        end_of_month = (
            (datetime(today.year, today.month + 1, 1) - timedelta(days=1)).date()
            if today.month < 12
            else datetime(today.year, 12, 31).date()
        )

        start_of_week = today - timedelta(days=today.weekday())
        end_of_week = start_of_week + timedelta(days=6)

        # Compile data
        data = {
            "week": [
                {
                    "name": staff.user.get_full_name(),  # Using full name instead of ID
                    "total": Appointment.objects.filter(
                        staff=staff.id,
                        appointment_date__range=[start_of_week, end_of_week],
                    ).count(),
                }
                for staff in Staff.objects.filter(user__company=request.user.company)
            ],
            "month": [
                {
                    "name": staff.user.get_full_name(),  # Using full name instead of ID
                    "total": Appointment.objects.filter(
                        staff=staff.id,
                        appointment_date__range=[start_of_month, end_of_month],
                    ).count(),
                }
                for staff in Staff.objects.filter(user__company=request.user.company)
            ],
            "year": [
                {
                    "name": staff.user.get_full_name(),  # Using full name instead of ID
                    "total": Appointment.objects.filter(
                        staff=staff.id,
                        appointment_date__range=[start_of_year, end_of_year],
                    ).count(),
                }
                for staff in Staff.objects.filter(user__company=request.user.company)
            ],
        }

        return Response(data, status=status.HTTP_200_OK)


# class AppointmentStatsByWeek(APIView):
#     permission_classes = [IsAdminOrStaff,]

#     def get(self,request,weekstart):
#         format_string = "%Y-%m-%d"
#         try:
#             is_valid = datetime.strptime(weekstart, format_string)
#             if not is_valid:
#                 return Response({"weekstart": "Invalid date"}, status = status.HTTP_400_BAD_REQUEST)
#         except:
#             return Response({"weekstart": "Invalid date"}, status = status.HTTP_400_BAD_REQUEST)
#         start_of_week = datetime.strptime(weekstart, format_string)
#         if not start_of_week.weekday() == 0:
#             return Response({"weekstart": "date is not start of the week (Monday)"}, status = status.HTTP_400_BAD_REQUEST)
#         end_of_week = start_of_week + timedelta(days=6)
#         data = [{"id": staff.id, "total": Appointment.objects.filter(staff = staff.id, appointment_date__range=[start_of_week, end_of_week]).count()} for staff in  Staff.objects.filter(user__company = request.user.company)]
#         return Response(data, status = status.HTTP_200_OK)


class AppointmentStatsByWeek(APIView):
    permission_classes = [
        IsAdminOrStaff,
    ]

    def get(self, request, weekstart):
        format_string = "%Y-%m-%d"
        try:
            is_valid = datetime.strptime(weekstart, format_string)
            if not is_valid:
                return Response(
                    {"weekstart": "Invalid date"}, status=status.HTTP_400_BAD_REQUEST
                )
        except ValueError:
            return Response(
                {"weekstart": "Invalid date"}, status=status.HTTP_400_BAD_REQUEST
            )

        start_of_week = datetime.strptime(weekstart, format_string)
        if start_of_week.weekday() != 0:  # Ensure it's a Monday
            return Response(
                {"weekstart": "Date is not the start of the week (Monday)"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        end_of_week = start_of_week + timedelta(days=6)

        data = [
            {
                "name": staff.user.get_full_name(),  # Assuming Staff is related to User and User has a method get_full_name()
                "total": Appointment.objects.filter(
                    staff=staff.id, appointment_date__range=[start_of_week, end_of_week]
                ).count(),
            }
            for staff in Staff.objects.filter(user__company=request.user.company)
        ]

        return Response(data, status=status.HTTP_200_OK)


class PatientMissedAppoinmentByWeek(APIView):
    permission_classes = [
        IsAdminOrStaff,
    ]

    def get(self, request, weekstart):
        format_string = "%Y-%m-%d"
        try:
            is_valid = datetime.strptime(weekstart, format_string)
            if not is_valid:
                return Response(
                    {"weekstart": "Invalid date"}, status=status.HTTP_400_BAD_REQUEST
                )
        except ValueError:
            return Response(
                {"weekstart": "Invalid date"}, status=status.HTTP_400_BAD_REQUEST
            )

        start_of_week = datetime.strptime(weekstart, format_string)
        if start_of_week.weekday() != 0:  # Ensure it's a Monday
            return Response(
                {"weekstart": "Date is not the start of the week (Monday)"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        end_of_week = start_of_week + timedelta(days=6)
        now = timezone.now()

        data = [
            {
                "name": client.user.get_full_name(),  # Assuming Staff is related to User and User has a method get_full_name()
                "total": Appointment.objects.filter(
                    client=client.id,
                    appointment_date__range=[start_of_week, end_of_week],
                    patient_arival=False,
                    appointment_date__lte=now.date(),  # Ensure the appointment is before today
                    appointment_start__lte=now.time(),  # Check if the appointment time is before current time
                ).count(),
            }
            for client in Client.objects.filter(
                assigned_staff__user__company=request.user.company
            )
        ]

        return Response(data, status=status.HTTP_200_OK)


class PatientMissedAppointmentStatsByMonth(APIView):
    permission_classes = [
        IsAdminOrStaff,
    ]

    def get(self, request, month_string):
        if not is_valid_date(month_string):
            return Response(
                {"month": "Invalid date"}, status=status.HTTP_400_BAD_REQUEST
            )

        month, year = month_string.split("-")
        month = int(month)
        year = int(year)

        start_of_month = datetime(year, month, 1).date()
        end_of_month = (
            (datetime(year, month + 1, 1) - timedelta(days=1)).date()
            if month < 12
            else datetime(year, 12, 31).date()
        )
        now = timezone.now()

        data = [
            {
                "name": client.user.get_full_name(),  # Assuming Staff is related to User and User has a method get_full_name()
                "total": Appointment.objects.filter(
                    client=client.id,
                    appointment_date__range=[start_of_month, end_of_month],
                    patient_arival=False,
                    appointment_date__lte=now.date(),  # Ensure the appointment is before today
                    appointment_start__lte=now.time(),  # Check if the appointment time is before current time
                ).count(),
            }
            for client in Client.objects.filter(
                assigned_staff__user__company=request.user.company
            )
        ]

        return Response(data, status=status.HTTP_200_OK)


class PatientMissedAppointmentStatsByYear(APIView):
    permission_classes = [
        IsAdminOrStaff,
    ]

    def get(self, request, year):
        start_of_year = datetime(year, 1, 1).date()
        end_of_year = datetime(year, 12, 31).date()
        now = timezone.now()

        data = [
            {
                "name": client.user.get_full_name(),  # Assuming Staff is related to User and User has a method get_full_name()
                "total": Appointment.objects.filter(
                    client=client.id,
                    appointment_date__range=[start_of_year, end_of_year],
                    patient_arival=False,
                    appointment_date__lte=now.date(),  # Ensure the appointment is before today
                    appointment_start__lte=now.time(),  # Check if the appointment time is before current time
                ).count(),
            }
            for client in Client.objects.filter(
                assigned_staff__user__company=request.user.company
            )
        ]

        return Response(data, status=status.HTTP_200_OK)


# class AppointmentStatsByMonth(APIView):
#     permission_classes = [IsAdminOrStaff,]

#     def get(self,request,month_string):
#         if not is_valid_date(month_string):
#             return Response({"month":"Invalid date"}, status = status.HTTP_400_BAD_REQUEST)
#         month, year = month_string.split('-')
#         month = int(month)
#         year = int(year)
#         start_of_month = datetime(year, month, 1).date()
#         end_of_month = (datetime(year, month + 1, 1) - timedelta(days=1)).date() if month < 12 else datetime(year, 12, 31).date()
#         data = [{"id": staff.id, "total": Appointment.objects.filter(staff = staff.id, appointment_date__range=[start_of_month, end_of_month]).count()} for staff in  Staff.objects.filter(user__company = request.user.company)]
#         # data = [{"name": staff.name, "total": Appointment.objects.filter(staff=staff.id, appointment_date__range=[start_of_month, end_of_month]).count()} for staff in Staff.objects.filter(user__company=request.user.company)]
#         return Response(data, status = status.HTTP_200_OK)
# class AppointmentStatsByMonth(APIView):
#     permission_classes = [IsAdminOrStaff,]

#     def get(self, request, month_string):
#         if not is_valid_date(month_string):
#             return Response({"month":"Invalid date"}, status = status.HTTP_400_BAD_REQUEST)
#         month, year = month_string.split('-')
#         month = int(month)
#         year = int(year)
#         start_of_month = datetime(year, month, 1).date()
#         end_of_month = (datetime(year, month + 1, 1) - timedelta(days=1)).date() if month < 12 else datetime(year, 12, 31).date()

#         # Modified to return staff name instead of id
#         data = [{"name": staff.name, "total": Appointment.objects.filter(staff=staff.id, appointment_date__range=[start_of_month, end_of_month]).count()} for staff in Staff.objects.filter(user__company=request.user.company)]


#         return Response(data, status=status.HTTP_200_OK)
######################################################################################
class AppointmentStatsByMonth(APIView):
    permission_classes = [
        IsAdminOrStaff,
    ]

    def get(self, request, month_string):
        if not is_valid_date(month_string):
            return Response(
                {"month": "Invalid date"}, status=status.HTTP_400_BAD_REQUEST
            )

        month, year = month_string.split("-")
        month = int(month)
        year = int(year)

        start_of_month = datetime(year, month, 1).date()
        end_of_month = (
            (datetime(year, month + 1, 1) - timedelta(days=1)).date()
            if month < 12
            else datetime(year, 12, 31).date()
        )

        data = [
            {
                "name": staff.user.get_full_name(),  # Assuming Staff is related to User and User has a method get_full_name()
                "total": Appointment.objects.filter(
                    staff=staff.id,
                    appointment_date__range=[start_of_month, end_of_month],
                ).count(),
            }
            for staff in Staff.objects.filter(user__company=request.user.company)
        ]

        return Response(data, status=status.HTTP_200_OK)


# class AppointmentStatsByYear(APIView):
#     permission_classes = [IsAdminOrStaff,]

#     def get(self,request,year):
#         if not re.match(r'^\d{4}$', str(year)):
#             return Response({"year":"Invalid date"}, status = status.HTTP_400_BAD_REQUEST)
#         start_of_year = datetime(year, 1, 1).date()
#         end_of_year = datetime(year, 12, 31).date()
#         data = [{"id": staff.id, "total": Appointment.objects.filter(staff = staff.id, appointment_date__range=[start_of_year, end_of_year]).count()} for staff in  Staff.objects.filter(user__company = request.user.company)]
#         return Response(data, status = status.HTTP_200_OK)


class AppointmentStatsByYear(APIView):
    permission_classes = [
        IsAdminOrStaff,
    ]

    def get(self, request, year):
        print(year)
        if not re.match(r"^\d{4}$", str(year)):
            return Response(
                {"year": "Invalid date"}, status=status.HTTP_400_BAD_REQUEST
            )

        start_of_year = datetime(year, 1, 1).date()
        end_of_year = datetime(year, 12, 31).date()
        print(start_of_year, end_of_year)
        data = [
            {
                "name": staff.user.get_full_name(),  # Assuming Staff is related to User and User has a method get_full_name()
                "total": Appointment.objects.filter(
                    staff=staff.id, appointment_date__range=[start_of_year, end_of_year]
                ).count(),
            }
            for staff in Staff.objects.filter(user__company=request.user.company)
        ]

        return Response(data, status=status.HTTP_200_OK)


class GeneralAssignedClientStats(APIView):
    permission_classes = [
        IsAdminOrStaff,
    ]

    def get(self, request):
        today = now().date()
        start_of_year = datetime(today.year, 1, 1).date()
        end_of_year = datetime(today.year, 12, 31).date()
        start_of_month = datetime(today.year, today.month, 1).date()
        end_of_month = (
            (datetime(today.year, today.month + 1, 1) - timedelta(days=1)).date()
            if today.month < 12
            else datetime(today.year, 12, 31).date()
        )
        start_of_week = today - timedelta(days=today.weekday())
        end_of_week = start_of_week + timedelta(days=6)
        data = {
            "week": [
                {
                    "id": staff.id,
                    "total": Client.objects.filter(
                        assigned_staff=staff.id,
                        user__date_joined__date__range=[start_of_week, end_of_week],
                    ).count(),
                }
                for staff in Staff.objects.filter(user__company=request.user.company)
            ],
            "month": [
                {
                    "id": staff.id,
                    "total": Client.objects.filter(
                        assigned_staff=staff.id,
                        user__date_joined__date__range=[start_of_month, end_of_month],
                    ).count(),
                }
                for staff in Staff.objects.filter(user__company=request.user.company)
            ],
            "year": [
                {
                    "id": staff.id,
                    "total": Client.objects.filter(
                        assigned_staff=staff.id,
                        user__date_joined__date__range=[start_of_year, end_of_year],
                    ).count(),
                }
                for staff in Staff.objects.filter(user__company=request.user.company)
            ],
        }
        return Response(data, status=status.HTTP_200_OK)


# class AssignedClientStatsByWeek(APIView):
#     permission_classes = [IsAdminOrStaff,]

#     def get(self,request,weekstart):
#         format_string = "%Y-%m-%d"
#         try:
#             is_valid = datetime.strptime(weekstart, format_string)
#             if not is_valid:
#                 return Response({"weekstart": "Invalid date"}, status = status.HTTP_400_BAD_REQUEST)
#         except:
#             return Response({"weekstart": "Invalid date"}, status = status.HTTP_400_BAD_REQUEST)
#         start_of_week = datetime.strptime(weekstart, format_string)
#         if not start_of_week.weekday() == 0:
#             return Response({"weekstart": "date is not start of the week (Monday)"}, status = status.HTTP_400_BAD_REQUEST)
#         end_of_week = start_of_week + timedelta(days=6)
#         data = [{"id": staff.id, "total": Client.objects.filter(assigned_staff = staff.id, user__date_joined__date__range=[start_of_week, end_of_week]).count()} for staff in  Staff.objects.filter(user__company = request.user.company)]
#         return Response(data, status = status.HTTP_200_OK)


class AssignedClientStatsByWeek(APIView):
    permission_classes = [
        IsAdminOrStaff,
    ]

    def get(self, request, weekstart):
        format_string = "%Y-%m-%d"
        try:
            is_valid = datetime.strptime(weekstart, format_string)
            if not is_valid:
                return Response(
                    {"weekstart": "Invalid date"}, status=status.HTTP_400_BAD_REQUEST
                )
        except ValueError:
            return Response(
                {"weekstart": "Invalid date"}, status=status.HTTP_400_BAD_REQUEST
            )

        start_of_week = datetime.strptime(weekstart, format_string)
        if start_of_week.weekday() != 0:  # Ensure it's a Monday
            return Response(
                {"weekstart": "Date is not the start of the week (Monday)"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        end_of_week = start_of_week + timedelta(days=6)

        data = [
            {
                "name": staff.user.get_full_name(),  # Using full name instead of ID
                "total": Client.objects.filter(
                    assigned_staff=staff.id,
                    user__date_joined__date__range=[start_of_week, end_of_week],
                ).count(),
            }
            for staff in Staff.objects.filter(user__company=request.user.company)
        ]

        return Response(data, status=status.HTTP_200_OK)


# class AssignedClientStatsByMonth(APIView):
#     permission_classes = [IsAdminOrStaff,]

#     def get(self,request,month_string):
#         if not is_valid_date(month_string):
#             return Response({"month":"Invalid date"}, status = status.HTTP_400_BAD_REQUEST)
#         month, year = month_string.split('-')
#         month = int(month)
#         year = int(year)
#         start_of_month = datetime(year, month, 1).date()
#         end_of_month = (datetime(year, month + 1, 1) - timedelta(days=1)).date() if month < 12 else datetime(year, 12, 31).date()
#         data = [{"id": staff.id, "total": Client.objects.filter(assigned_staff = staff.id, user__date_joined__date__range=[start_of_month, end_of_month]).count()} for staff in  Staff.objects.filter(user__company = request.user.company)]
#         return Response(data, status = status.HTTP_200_OK)


class AssignedClientStatsByMonth(APIView):
    permission_classes = [
        IsAdminOrStaff,
    ]

    def get(self, request, month_string):
        if not is_valid_date(month_string):
            return Response(
                {"month": "Invalid date"}, status=status.HTTP_400_BAD_REQUEST
            )

        month, year = month_string.split("-")
        month = int(month)
        year = int(year)

        start_of_month = datetime(year, month, 1).date()
        end_of_month = (
            (datetime(year, month + 1, 1) - timedelta(days=1)).date()
            if month < 12
            else datetime(year, 12, 31).date()
        )

        data = [
            {
                "name": staff.user.get_full_name(),  # Using full name instead of ID
                "total": Client.objects.filter(
                    assigned_staff=staff.id,
                    user__date_joined__date__range=[start_of_month, end_of_month],
                ).count(),
            }
            for staff in Staff.objects.filter(user__company=request.user.company)
        ]

        return Response(data, status=status.HTTP_200_OK)


# class AssignedClientStatsByYear(APIView):
#     permission_classes = [IsAdminOrStaff,]

#     def get(self,request,year):
#         start_of_year = datetime(year, 1, 1).date()
#         end_of_year = datetime(year, 12, 31).date()
#         data = [{"id": staff.id, "total": Client.objects.filter(assigned_staff = staff.id, user__date_joined__date__range=[start_of_year, end_of_year]).count()} for staff in  Staff.objects.filter(user__company = request.user.company)]
#         return Response(data, status = status.HTTP_200_OK)


class AssignedClientStatsByYear(APIView):
    permission_classes = [
        IsAdminOrStaff,
    ]

    def get(self, request, year):
        start_of_year = datetime(year, 1, 1).date()
        end_of_year = datetime(year, 12, 31).date()

        data = [
            {
                "name": staff.user.get_full_name(),  # Using full name instead of ID
                "total": Client.objects.filter(
                    assigned_staff=staff.id,
                    user__date_joined__date__range=[start_of_year, end_of_year],
                ).count(),
            }
            for staff in Staff.objects.filter(user__company=request.user.company)
        ]

        return Response(data, status=status.HTTP_200_OK)


######################################################################
from datetime import datetime
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from dateutil.relativedelta import relativedelta
import calendar


class ClientAgeStatsByYear(APIView):
    permission_classes = [IsAdminOrStaff]

    def get(self, request, year):
        # Validate the year format
        if not re.match(r"^\d{4}$", str(year)):
            return Response(
                {"year": "Invalid date format. Expected YYYY"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Make datetime objects timezone-aware
        start_of_year = timezone.make_aware(datetime(int(year), 1, 1))
        end_of_year = timezone.make_aware(datetime(int(year), 12, 31, 23, 59, 59))

        # Fetch all clients within the year
        clients = Client.objects.filter(
            user__company=request.user.company,
            user__date_joined__range=[start_of_year, end_of_year],
        )

        # Initialize the age ranges
        age_ranges = {
            "<18": 0,
            "18-24": 0,
            "25-34": 0,
            "35-44": 0,
            "45-54": 0,
            "55-64": 0,
            "65+": 0,
        }

        # Calculate and categorize the ages
        for client in clients:
            if client.user.date_birth:
                age = relativedelta(timezone.now().date(), client.user.date_birth).years
                if age < 18:
                    age_ranges["<18"] += 1
                elif 18 <= age <= 24:
                    age_ranges["18-24"] += 1
                elif 25 <= age <= 34:
                    age_ranges["25-34"] += 1
                elif 35 <= age <= 44:
                    age_ranges["35-44"] += 1
                elif 45 <= age <= 54:
                    age_ranges["45-54"] += 1
                elif 55 <= age <= 64:
                    age_ranges["55-64"] += 1
                else:
                    age_ranges["65+"] += 1

        data = {"total": clients.count(), "age_distribution": age_ranges}

        return Response(data, status=status.HTTP_200_OK)


class ClientAgeStatsByMonth(APIView):
    permission_classes = [IsAdminOrStaff]

    def get(self, request, month_string):
        # Validate the month format
        try:
            month, year = map(int, month_string.split("-"))
            if month < 1 or month > 12:
                raise ValueError("Invalid month")
        except ValueError:
            return Response(
                {"month": "Invalid month format. Expected MM-YYYY with month in 1-12"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        start_of_month = datetime(year, month, 1).date()
        last_day = calendar.monthrange(year, month)[1]
        end_of_month = datetime(year, month, last_day).date()

        # Fetch clients who joined in the given month
        clients = Client.objects.filter(
            user__company=request.user.company,
            user__date_joined__range=[start_of_month, end_of_month],
        )

        # Initialize the age ranges
        age_ranges = {
            "<18": 0,
            "18-24": 0,
            "25-34": 0,
            "35-44": 0,
            "45-54": 0,
            "55-64": 0,
            "65+": 0,
        }

        # Calculate and categorize the ages
        for client in clients:
            if client.user.date_birth:
                age = relativedelta(timezone.now().date(), client.user.date_birth).years
                if age < 18:
                    age_ranges["<18"] += 1
                elif 18 <= age <= 24:
                    age_ranges["18-24"] += 1
                elif 25 <= age <= 34:
                    age_ranges["25-34"] += 1
                elif 35 <= age <= 44:
                    age_ranges["35-44"] += 1
                elif 45 <= age <= 54:
                    age_ranges["45-54"] += 1
                elif 55 <= age <= 64:
                    age_ranges["55-64"] += 1
                else:
                    age_ranges["65+"] += 1

        data = {"total": clients.count(), "age_distribution": age_ranges}

        return Response(data, status=status.HTTP_200_OK)


class ClientAgeStatsByWeek(APIView):
    permission_classes = [
        IsAdminOrStaff,
    ]

    def get(self, request, week_start_date):
        # Validate the date format
        try:
            start_of_week = datetime.strptime(week_start_date, "%Y-%m-%d").date()
        except ValueError:
            return Response(
                {"week_start_date": "Invalid date format. Expected YYYY-MM-DD"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if start_of_week.weekday() != 0:
            return Response(
                {"week_start_date": "The provided date is not a Monday"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        end_of_week = start_of_week + timedelta(days=6)

        # Fetch clients who joined in the given week
        clients = Client.objects.filter(
            user__company=request.user.company,
            user__date_joined__range=[start_of_week, end_of_week],
        )

        # Initialize the age ranges
        age_ranges = {
            "<18": 0,
            "18-24": 0,
            "25-34": 0,
            "35-44": 0,
            "45-54": 0,
            "55-64": 0,
            "65+": 0,
        }

        # Calculate and categorize the ages
        for client in clients:
            if client.user.date_birth:
                age = relativedelta(datetime.now().date(), client.user.date_birth).years
                if age < 18:
                    age_ranges["<18"] += 1
                elif 18 <= age <= 24:
                    age_ranges["18-24"] += 1
                elif 25 <= age <= 34:
                    age_ranges["25-34"] += 1
                elif 35 <= age <= 44:
                    age_ranges["35-44"] += 1
                elif 45 <= age <= 54:
                    age_ranges["45-54"] += 1
                elif 55 <= age <= 64:
                    age_ranges["55-64"] += 1
                else:
                    age_ranges["65+"] += 1

        data = {"total": clients.count(), "age_distribution": age_ranges}

        return Response(data, status=status.HTTP_200_OK)


######################################################################


# class AppointmentSendInfoStatsView(APIView):

#     def get(self, request, *args, **kwargs):
#         year = request.query_params.get('year', None)
#         month = request.query_params.get('month', None)
#         week = request.query_params.get('week', None)

#         # If no year is provided, use the current year
#         if year is None:
#             year = timezone.now().year
#         else:
#             year = int(year)

#         if month is not None:
#             month = int(month)
#             sends_count = AppointmentSendInfo.objects.filter(send_at__year=year, send_at__month=month).count()
#             period = f"Year: {year}, Month: {month}"
#         elif week is not None:
#             week = int(week)
#             sends_count = AppointmentSendInfo.objects.filter(send_at__year=year, send_at__week=week).count()
#             period = f"Year: {year}, Week: {week}"
#         else:
#             sends_count = AppointmentSendInfo.objects.filter(send_at__year=year).count()
#             period = f"Year: {year}"

#         data = {
#             'period': period,
#             'sends_count': sends_count
#         }
#         return Response(data, status=status.HTTP_200_OK)


class AppointmentSendInfoStatsView(APIView):

    def get(self, request, *args, **kwargs):
        year = request.query_params.get("year", None)
        month = request.query_params.get("month", None)
        week = request.query_params.get("week", None)

        if year is None:
            year = timezone.now().year
        else:
            year = int(year)

        data = []

        for i in range(6):
            current_year = year + i

            if i == 0:
                if month is not None:
                    month = int(month)
                    sends_count = AppointmentSendInfo.objects.filter(
                        send_at__year=current_year, send_at__month=month
                    ).count()
                    period = f"{month}-{current_year}"
                elif week is not None:
                    week = int(week)
                    sends_count = AppointmentSendInfo.objects.filter(
                        send_at__year=current_year, send_at__week=week
                    ).count()
                    period = f"{week}-{current_year}"
                else:
                    sends_count = AppointmentSendInfo.objects.filter(
                        send_at__year=current_year
                    ).count()
                    period = current_year
            else:
                sends_count = 0
                if month is not None:
                    period = f"{month}-{current_year}"
                elif week is not None:
                    period = f"{week}-{current_year}"
                else:
                    period = current_year

            data.append({"period": period, "sends_count": sends_count})

        return Response(data, status=status.HTTP_200_OK)


class CreateClientCommentView(APIView):
    permission_classes = [
        IsAuthenticated
    ]  # Only authenticated users can create comments

    def get(self, request, *args, **kwargs):
        user_id = request.query_params.get("user_id")
        comments = ClientComment.objects.filter(user__id=user_id)
        serializer = ClientCommentSerializer(comments, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    def post(self, request, *args, **kwargs):
        serializer = ClientCommentSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save(
                user=request.user
            )  # Automatically set the user to the current authenticated user
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class NoteCreate(APIView):
    # queryset = Note.objects.all()
    serializer_class = NoteSerializer
    permission_classes = [IsAdminUser]  # Ensure only authenticated users can access

    def get(self, request, **args):
        queryset = Note.objects.filter(author=request.user)
        serializer = NoteSerializer(queryset, many=True)
        return Response(serializer.data)

    def post(self, request, *args, **kwargs):
        serializer = self.serializer_class(data=request.data)
        if serializer.is_valid():
            serializer.save(
                author=self.request.user
            )  # Automatically set the author as the logged-in user
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def put(self, request, *args, **kwargs):
        # Fetch the note to update
        note_id = request.data.get("note_id")
        note = get_object_or_404(Note, id=note_id, author=request.user)

        # Validate and update the note data
        serializer = self.serializer_class(
            note, data=request.data, partial=True
        )  # `partial=True` allows partial updates
        if serializer.is_valid():
            serializer.save()  # Save the updated note
            return Response(serializer.data, status=status.HTTP_200_OK)

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def delete(self, request, *args, **kwargs):
        note_id = request.query_params.get("note_id")
        print(note_id)
        note = get_object_or_404(Note, id=note_id, author=request.user)
        note.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


class TaskManagement(APIView):
    permission_classes = [
        IsAdminOrStaff,
    ]
    serializer_class = TaskSerializer

    def get(self, request, *args, **kwargs):
        if request.user.user_role == "STAFF":
            tasks = Task.objects.filter(created_by=request.user)
            serializer = TaskSerializer(
                tasks, many=True
            )  # Use 'many=True' to serialize a queryset
            return Response(serializer.data, status=status.HTTP_200_OK)
        if request.user.user_role == "ADMIN":
            tasks = Task.objects.filter(assigned_admins=request.user)
            serializer = TaskSerializer(
                tasks, many=True
            )  # Use 'many=True' to serialize a queryset
            return Response(serializer.data, status=status.HTTP_200_OK)

    def post(self, request, *args, **kwargs):
        serializer = self.serializer_class(data=request.data)
        if serializer.is_valid():
            # Save the task and set the 'created_by' field to the current user
            task = serializer.save(created_by=request.user)

            # Handle the many-to-many field (assigned_admins)
            assigned_admins = serializer.validated_data.pop("assigned_admins", [])
            task.assigned_admins.set(
                assigned_admins
            )  # Set the assigned admins to the task

            # Check if send_email is true and send emails to assigned admins
            if request.data.get("send_email", False):
                assigned_admins = request.data.get("assigned_admins", [])
                for admin_id in assigned_admins:
                    # Fetch the admin user based on ID
                    try:
                        admin = CustomUser.objects.get(id=admin_id)

                        # Correctly set the subject as a string
                        subject = "New Task Assigned"
                        message = (
                            f'A new task "{task.description}" has been assigned to you.'
                        )

                        email = EmailMessage(subject, message, to=[admin.email])
                        email.content_subtype = (
                            "html"  # Specify the content type as HTML
                        )
                        email.send()
                    except CustomUser.DoesNotExist:
                        print(f"Admin with id {admin_id} does not exist.")
                    except Exception as e:
                        print(f"Error sending email: {e}")

            return Response(serializer.data, status=status.HTTP_201_CREATED)

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def put(self, request, *args, **kwargs):
        try:
            task_id = request.data.get("task_id")
            new_status = request.data.get("status")

            # Fetch the task by ID
            try:
                task = Task.objects.get(id=task_id)
            except Task.DoesNotExist:
                return Response(
                    {"error": "Task not found."}, status=status.HTTP_404_NOT_FOUND
                )

            # Validate the status (optional, based on your application's rules)
            if new_status not in ["pending", "in_progress", "completed", "cancelled"]:
                return Response(
                    {"error": "Invalid status value."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            # Update the status
            task.status = new_status
            task.save()

            return Response(
                {"message": "Task status updated successfully."},
                status=status.HTTP_200_OK,
            )

        except Exception as e:
            return Response(
                {"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class SendTaskEmailAPIView(APIView):
    permission_classes = [
        IsAdminOrStaff,
    ]

    def post(self, request, *args, **kwargs):
        try:
            task_id = request.data.get("task_id", None)
            task = Task.objects.get(id=task_id)
            print(task, task_id)

            # Iterate through assigned_admins and send emails
            for admin in task.assigned_admins.all():  # Use parentheses to call all()
                try:
                    # Fetch the admin user based on ID
                    admin_user = CustomUser.objects.get(
                        id=admin.id
                    )  # Access the user object

                    # Set the subject and message for the email
                    subject = f"Task Notification: {task_id}"
                    message = (
                        f"A new task has been assigned to you:\n\n"
                        f"Task ID: {task_id}\n"
                        f"Description: {task.description}\n"
                        f"Status: {task.status}\n"
                        f"Start Date: {task.start_date}\n"
                        f"Expected Completion Date: {task.expected_completion_date}\n"
                    )

                    # Create and send the email
                    email = EmailMessage(subject, message, to=[admin_user.email])
                    email.content_subtype = "html"  # Specify the content type as HTML
                    email.send()

                except CustomUser.DoesNotExist:
                    print(f"Admin with id {admin.id} does not exist.")
                except Exception as e:
                    print(f"Error sending email: {e}")

            return Response(
                {"message": "Email sent successfully!"}, status=status.HTTP_200_OK
            )

        except Task.DoesNotExist:
            return Response(
                {"error": "Task not found."}, status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response(
                {"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class TaskCommentView(APIView):
    permission_classes = [
        IsAdminOrStaff,
    ]
    serializer_class = TaskCommentSerializer

    def get(self, request, *args, **kwargs):
        task_id = request.query_params.get("task_id")
        if task_id is None:
            return Response(
                {"error": "Task ID not provided"}, status=status.HTTP_400_BAD_REQUEST
            )
        print(task_id)
        try:
            # Fetch the task comments using the task_id
            task = Task.objects.get(id=task_id)
            task_comments = TaskComment.objects.filter(task=task)
            print(task_comments)
            serializer = TaskCommentSerializer(task_comments, many=True)
            print(serializer.data)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except Task.DoesNotExist:
            return Response(
                {"error": "Task not found"}, status=status.HTTP_404_NOT_FOUND
            )

    def post(self, request, *args, **kwargs):
        serializer = self.serializer_class(data=request.data)
        if serializer.is_valid():
            serializer.save(author=request.user)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


from django.http import HttpResponse
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.lib.utils import simpleSplit


class GenerateUserPDF(APIView):
    def get(self, request):
        try:
            user = CustomUser.objects.get(id=request.user.id)
        except CustomUser.DoesNotExist:
            return Response({"error": "User not found"}, status=404)

        # Create PDF response
        response = HttpResponse(content_type="application/pdf")
        response["Content-Disposition"] = f'attachment; filename="user_{user.id}.pdf"'

        # Initialize PDF canvas
        pdf = canvas.Canvas(response, pagesize=A4)
        width, height = A4  # Get the A4 page dimensions

        # Page Title
        pdf.setFont("Helvetica-Bold", 16)
        pdf.setFillColor(colors.darkblue)
        pdf.drawString(200, height - 80, "User Details Report")

        # Draw a line under title
        pdf.setStrokeColor(colors.black)
        pdf.setLineWidth(2)
        pdf.line(50, height - 90, width - 50, height - 90)

        # Set font for details
        pdf.setFont("Helvetica", 12)
        pdf.setFillColor(colors.black)

        # Start position for details
        y = height - 120
        line_spacing = 18  # Line height

        # User Details Dictionary
        user_details = {
            "User ID": user.id,
            "Full Name": f"{user.first_name} {user.last_name}",
            "Email": user.email,
            "Role": "Patient",
            "Company": user.company.name if user.company else "N/A",
            "Specializations": user.specializations or "N/A",
            "Department": user.department or "N/A",
            "City": user.city or "N/A",
            "Address": user.address or "N/A",
            "Postcode": user.postcode or "N/A",
            "Gender": user.get_gender_display() if user.gender else "N/A",
            "Phone": user.phone_number if user.phone_number else "N/A",
            "Date of Birth": (
                user.date_birth.strftime("%Y-%m-%d") if user.date_birth else "N/A"
            ),
            "GESY Number": user.gesy_number or "N/A",
            "ID Card": user.id_card or "N/A",
            "Color Code": user.color or "N/A",
        }

        # Draw the user details in key-value format
        pdf.setFont("Helvetica", 11)
        for key, value in user_details.items():
            pdf.setFillColor(colors.darkblue)  # Label color
            pdf.drawString(60, y, f"{key}:")
            pdf.setFillColor(colors.black)  # Value color
            pdf.drawString(180, y, str(value))

            y -= line_spacing  # Move to next line

            # Prevent writing off the page
            if y < 50:
                pdf.showPage()  # New page
                pdf.setFont("Helvetica", 12)
                y = height - 80  # Reset Y position

        # Save and close PDF
        pdf.showPage()
        pdf.save()

        return response
